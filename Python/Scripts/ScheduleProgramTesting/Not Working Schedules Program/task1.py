import openpyxl
fname = r'C:\Users\Yousef\Downloads\DevGroup-Task1-Dicts.xlsx'
wb = openpyxl.load_workbook(fname)
sheet = wb['Sheet1']
dict_Q={}
srv=0
max=sheet.cell(row=sheet.min_row+1,column =sheet.min_column+3).value
number=0
percentage=[]
y=sheet.min_row+1
for x in range(sheet.min_row+1,sheet.max_row+1):
    dict_Q[sheet.cell(row=x,column =sheet.min_column).value]=[sheet.cell(row=x,column =sheet.min_column+1).value,sheet.cell(row=x,column =sheet.min_column+2).value,sheet.cell(row=x,column =sheet.min_column+3).value]
    srv+=dict_Q[sheet.cell(row=x,column =sheet.min_column).value][0]
    if sheet.cell(row=x,column =sheet.min_column+3).value >max:
        max=sheet.cell(row=x,column =sheet.min_column+3).value
        number=x
while y+4 <= sheet.max_row:
    percentage += [abs(((sheet.cell(row=y+4,column =sheet.min_column+1).value - sheet.cell(row=y,column =sheet.min_column+1).value)/sheet.cell(row=y,column =sheet.min_column+1).value)*100)]
    y+=1
print("The total srv: ",srv)
print("The 'srv' value for the row which has the maximum 'EP' value: ",sheet.cell(row=number,column =sheet.min_column+1).value)
print("The percentage: ",percentage)
