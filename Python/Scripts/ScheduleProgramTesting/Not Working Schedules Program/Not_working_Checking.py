import openpyxl
import itertools
import os
from openpyxl import Workbook
book = Workbook()
sheet = book.active
fname = r'C:\Users\Yousef\Documents\Python\appending.xlsx'
wb = openpyxl.load_workbook(fname)
sheets = wb.sheetnames
LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
dict={"math tut":2,"edpt tut":1,"de tut":2,"csen tut":2,"sm tut":1,"elct tut":1,"phys tut":1,"math lecture":1,"math2 lecture":1,"production lecture":1,"cs lecture":1,"elct lecture":1,"physics lecture":1}
check={"math tut":0,"edpt tut":0,"de tut":0,"csen tut":0,"sm tut":0,"elct tut":0,"phys tut":0,"math lecture":0,"math2 lecture":0,"production lecture":0,"cs lecture":0,"elct lecture":0,"physics lecture":0}
def excel_style(row, col):
    """ Convert given row and column number to an Excel-style cell name. """
    result = []
    while col:
        col, rem = divmod(col-1, 26)
        result[:0] = LETTERS [rem]
    return ''.join(result) + str(row)# the cell name system for excel   #getting the name of a cell in the format 'A1'  # the cell name system for excel   #getting the name of a cell in the format 'A1'
def collecting_sheets_data():
    list=[]
    list_of_lists=[]
    for thesheet in sheets:
        sheet=wb[thesheet]
        list=[]
        for x in range(sheet.min_row,sheet.max_row+1):
            for y in range(sheet.min_column,sheet.max_column+1):
                list+=[sheet.cell(row=x,column =y).value]
        list_of_lists+=[list]
    return list_of_lists
def getting_final_lists():
    my_lists=collecting_sheets_data()
    list_without_duplicates=[]
    for i in my_lists:
        if i not in list_without_duplicates:
            list_without_duplicates.append(i)
    final_list=[]
    for list in list_without_duplicates:
        check={"math tut":0,"edpt tut":0,"de tut":0,"csen tut":0,"sm tut":0,"elct tut":0,"phys tut":0,"math lecture":0,"math2 lecture":0,"production lecture":0,"cs lecture":0,"elct lecture":0,"physics lecture":0}
        for slot in list:
            for key,value in check.items():
                if slot == None:
                    continue
                if slot.lower()==key.lower():
                    check[key]+=1
                    break
        if check == dict:
            final_list.append(list)
    return final_list
def appending_final_lists():
    sheet = book.active
    x=1
    y=1
    i=0
    for list in getting_final_lists():
        x=1
        y=1
        for slot in list:
            sheet[excel_style(y,x)]= slot
            if x >5:
                x=0
                y+=1
            x+=1
        i+=1
        if i<len(getting_final_lists()):
            sheet=book.create_sheet("sheet " +str(i))
appending_final_lists()

book.save(r'C:\Users\Yousef\Documents\Python\Your_Schedules.xlsx')
file = r'C:\Users\Yousef\Documents\Python\Your_Schedules.xlsx'
os.startfile(file)
