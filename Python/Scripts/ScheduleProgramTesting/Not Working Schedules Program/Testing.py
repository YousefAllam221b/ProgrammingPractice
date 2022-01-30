import itertools
import openpyxl
import random
from openpyxl import Workbook
book = Workbook()
sheet = book.active
fname = r'C:\Users\Yousef\Documents\Python\Schedule.xlsx'
wb = openpyxl.load_workbook(fname)
sheet1 = wb['Sheet1']
LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
weekdays=["saturday","sunday","monday","tuesday","wednesday","thursday"]
holidays=[]
math={"saturday":["1st"],"monday":["3rd","4th"]}
math2={"wednesday":["2nd","4th"],"thursday":["4th"]}
physics={"saturday":["2nd","3rd"],"sunday":["4th"],"monday":["2nd"],"wednesday":["2nd","4th"]}
cs={"tuesday":["3rd","4th"],"thursday":["2nd","3rd"]}
production={"monday":["1st","2nd"],"wednesday":["1st","4th"]}
elct={"saturday":["2nd","4th"],"sunday":["3rd","4th"],"thursday":["4th"]}
subjects=["math","pyhsics","cs","prodcution","elct"]

def Weekdays(weekdays):
    theWeekdays=[]
    holidays=[]
    flag=True
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            if sheet1.cell(row=x,column =y).value != None:
                flag = False
        if flag==True:
            holidays+=[sheet1.cell(row=x,column =sheet1.min_column).value]
            flag=False
        else:
            theWeekdays+=[sheet1.cell(row=x,column =sheet1.min_column).value]
        flag=True
    return theWeekdays,holidays #returns weekdays while the other returns holidays
def creatingTubles(subject):
    tubles=[]
    for key,value in subject.items():
        for val in value:
            tubles+=[(key,val)]
    return tubles #creates tubles of times for a subject
def emptySlot():
    horizontal=sheet1.min_row+1
    vertical=sheet1.min_column+1
    slots=[]
    slotsNames=[]
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            if len(Weekdays(weekdays)[1]) ==2:
                if Weekdays(weekdays)[1][0] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[1][1] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(horizontal,vertical)]
                            slotsNames+=[(sheet1.cell(row=x,column =sheet1.min_column).value,sheet1.cell(row=sheet1.min_row,column =y).value)]
            elif len(Weekdays(weekdays)[1]) ==1:
                if Weekdays(weekdays)[1][0] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(horizontal,vertical)]
                            slotsNames+=[(sheet1.cell(row=x,column =sheet1.min_column).value,sheet1.cell(row=sheet1.min_row,column =y).value)]
            elif len(Weekdays(weekdays)[1]) ==3:
                if Weekdays(weekdays)[1][0] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[1][1] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[1][2] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(horizontal,vertical)]
                            slotsNames+=[(sheet1.cell(row=x,column =sheet1.min_column).value,sheet1.cell(row=sheet1.min_row,column =y).value)]
            vertical+=1
        horizontal+=1
        vertical=sheet1.min_column+1
    return slotsNames,slots
def excel_style(row, col):
    """ Convert given row and column number to an Excel-style cell name. """
    result = []
    while col:
        col, rem = divmod(col-1, 26)
        result[:0] = LETTERS [rem]
    return ''.join(result) + str(row)# the cell name system for excel   #getting the name of a cell in the format 'A1'  # the cell name system for excel   #getting the name of a cell in the format 'A1'
def removeHolidayS(Asubject): #removes the holidays times from the subjects times
    dayoff= Weekdays(weekdays)[1]
    slots=creatingTubles(Asubject)
    final=[]
    for slot in slots:
        if slot in emptySlot()[0]:
            final+=[slot]
    return final
math_=removeHolidayS(math)
math2_=removeHolidayS(math2)
physics_=removeHolidayS(physics)
cs_=removeHolidayS(cs)
elct_=removeHolidayS(elct)
production_=removeHolidayS(production)

def permutations():
    empty2=["math","math2","physics","cs","production","elct"]
    empty=list(itertools.permutations(empty2))
    return empty
def fullSlot():
    horizontal=sheet1.min_row+1
    vertical=sheet1.min_column+1
    slots=[]
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            if sheet1.cell(row=x,column =y).value != None:
                slots+=[(horizontal,vertical)]
            vertical+=1
        horizontal+=1
        vertical=sheet1.min_column+1
    return slots
def copingOldSheet(the_sheet):
    i=0
    for x in range(sheet1.min_row,sheet1.max_row+1):
            for y in range(sheet1.min_column,sheet1.max_column+1):
                if sheet1.cell(row=x,column=y).value !=None:
                    the_sheet[excel_style(x,y)]=sheet1.cell(row=x,column=y).value
def getitmvalue(x):
    if x == 'math':
            return math_
    elif x == 'math2':
            return math2_
    elif x == 'physics':
            return physics_
    elif x == 'cs':
            return cs_
    elif x == 'production':
            return production_
    elif x == 'elct':
            return elct_


#math_=removeHolidayS(math)
#ath2_=removeHolidayS(math2)
#physics_=removeHolidayS(physics)
#cs_=removeHolidayS(cs)
#elct_=removeHolidayS(elct)
#production_=removeHolidayS(production)

def PuttingValues():
    sheet=book.active
    options=permutations()
    subjects=["math","math2","physics","cs","production","elct"]

    dict={'math':math_,'math2':math2_,'physics':physics_,'cs':cs_,'production':production_,'elct':elct_}
    empty=emptySlot()[0]
    emptyS=emptySlot()[1]
    empty_options=[]
    check_number={"math":0,"math2":0,"cs":0,"physics":0,"production":0,"elct":0}
    flag404=False
    counting_numbers={"math":0,"math2":0,"cs":0,"physics":0,"production":0,"elct":0}
    i=0
    appended_slots={"math":0,"math2":0,"cs":0,"physics":0,"production":0,"elct":0}
    counter=-1
    lists=[]
    output=[]
    assigned=[]
    for anOption in options:

        for subject in subjects:
            itmval = getitmvalue(subject)
            print("vlaues before:",itmval)
            print("dictitems",dict.items())
            for value in itmval:
                empty=emptySlot()[0]
                emptyS=emptySlot()[1]
                print(elct_)
                check_number={"math":0,"math2":0,"cs":0,"physics":0,"production":0,"elct":0}
                for row in sheet['A1:G37']:
                  for cell in row:
                    cell.value = None
                output=[]
                assigned=[]

                for subject_name in anOption:
                    if check_number[subject_name] ==1:
                        continue
                    if subject == subject_name:
                        for empty_slot in empty:
                            if empty_slot not in assigned:
                                print('hi',value,empty_slot,subject,subject_name)
                                if value == empty_slot:

                                    counter= empty.index(empty_slot)
                                    check_number[subject_name]+=1
                                    counting_numbers[subject_name]+=1
                                    output+=[(subject_name,empty_slot)]
                                    assigned+=[empty_slot]

                                    print('hiafter',value,empty_slot,subject,subject_name, empty_slot==("saturday","3rd"),subject_name=="elct")
                                    if subject_name =="elct":
                                        if empty_slot ==("saturday","3rd"):
                                            print("yes")
                                break
                    else:
                        for subject_slot in dict[subject_name]:
                            for empty_slot in empty:
                                if empty_slot not in assigned:
                                    if subject_slot == empty_slot:
                                        counter= empty.index(empty_slot)
                                        sheet[excel_style(emptyS[counter][0],emptyS[counter][1])]= subject_name + " lecture"
                                        check_number[subject_name]+=1
                                        counting_numbers[subject_name]+=1
                                        output+=[(subject_name,empty_slot)]
                                        assigned+=[empty_slot]
                                        break
                            if check_number[subject_name] == 1:
                                break
                flag=True

                if {'math': 1, 'math2': 1, 'cs': 1, 'physics': 1, 'production': 1, 'elct': 1} == check_number:
                    print(check_number)
                for item,itsValue in check_number.items():
                    if itsValue !=1:
                        flag=False
                        break

                if flag==True:
                    order=["math","math2","physics","cs","production","elct"]
                    for subject in order:
                        for tuple in output:
                            if tuple[0] == subject:
                                counter=order.index(subject)
                                other=output.index(tuple)
                                temp=output[counter]
                                output[counter]=tuple
                                output[other]=temp


                    if output in lists:
                        pass
                    else:
                        print("**********************77")
                        lists+=[output]
                        print(output)

                        print("**********************88")






PuttingValues()


book.save(r'C:\Users\Yousef\Documents\Python\appending.xlsx')
