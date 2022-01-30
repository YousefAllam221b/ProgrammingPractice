import openpyxl
import random
from openpyxl import Workbook
book = Workbook()
sheet = book.active
fname = r'C:\Users\Yousef\Documents\Python\Schedule.xlsx'
wb = openpyxl.load_workbook(fname)
sheet1 = wb['Sheet1']
LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
weekdays=["saturday","sunday","monday","tuesday","wednesday","thursday"]
holidays=[]
math={"saturday":["1st"],"monday":["3rd","4th"]}
math2={"wednesday":["2nd","4th"],"thursday":["4th"]}
physics={"saturday":["2nd","3rd"],"sunday":["4th"],"monday":["2nd"],"wednesday":["2nd","4th"]}
cs={"tuesday":["3rd"],"thursday":["2nd","3rd"]}
production={"monday":["1st","2nd"],"wednesday":["1st","4th"]}
elct={"saturday":["2nd","4th"],"sunday":["3rd","4th"],"thursday":["4th"]}
subjects=["math","pyhsics","cs","prodcution","elct"]
def Weekdays(weekdays):
    holidays=[]
    flag=True
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            if sheet1.cell(row=x,column =y).value != None:
                flag = False
        if flag==True:
            holidays+=[sheet1.cell(row=x,column =sheet1.min_column).value]
            flag=False
        flag=True
    return holidays     #return a list of HOLIDAYS  #creates a list of the holidays
def Weekdays2(weekdays):
    holidays=[]
    flag=True
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            if sheet1.cell(row=x,column =y).value != None:
                flag = False
        if flag==True:
            weekdays.remove(sheet1.cell(row=x,column =sheet1.min_column).value)
            flag=False
        flag=True
    return weekdays #returns weekdays while the other returns holidays
def creatingTubles(subject):
    tubles=[]
    for key,value in subject.items():
        for val in value:
            tubles+=[(key,val)]
    return tubles #creates tubles of times for a subject
def emptySlot2():
    horizontal=sheet1.min_row+1
    vertical=sheet1.min_column+1
    slots=[]
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            if len(Weekdays(weekdays)) ==2:
                if Weekdays(weekdays)[0] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[1] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(sheet1.cell(row=x,column =sheet1.min_column).value,sheet1.cell(row=sheet1.min_row,column =y).value)]
            elif len(Weekdays(weekdays)) ==1:
                if Weekdays(weekdays)[0] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(horizontal,vertical)]
            elif len(Weekdays(weekdays)) ==3:
                if Weekdays(weekdays)[0] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[1] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[2] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(horizontal,vertical)]
            vertical+=1
        horizontal+=1
        vertical=sheet1.min_column+1
    return slots    #getting empty slots from the Schedule before statring  #for the empty slots of the original schedule
def emptySlot():
    horizontal=sheet1.min_row+1
    vertical=sheet1.min_column+1
    slots=[]
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            if len(Weekdays(weekdays)) ==2:
                if Weekdays(weekdays)[0] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[1] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(horizontal,vertical)]
            elif len(Weekdays(weekdays)) ==1:
                if Weekdays(weekdays)[0] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(horizontal,vertical)]
            elif len(Weekdays(weekdays)) ==3:
                if Weekdays(weekdays)[0] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[1] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[2] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(horizontal,vertical)]
            vertical+=1
        horizontal+=1
        vertical=sheet1.min_column+1
    return slots
def converSlotToCell():
    names=[]
    for x in emptySlot():
        names+=[excel_style(x[0],x[1])]
    return names    #convert a slot to a name of cell using coordinates #convert cell coordinates to names
def coordinate(self):
        """This cell's coordinate (ex. 'A5')"""
        col = get_column_letter(self.column)
        return f"{col}{self.row}"   # not working yet
def excel_style(row, col):
    """ Convert given row and column number to an Excel-style cell name. """
    result = []
    while col:
        col, rem = divmod(col-1, 26)
        result[:0] = LETTERS [rem]
    return ''.join(result) + str(row)# the cell name system for excel   #getting the name of a cell in the format 'A1'  # the cell name system for excel   #getting the name of a cell in the format 'A1'
def removeHolidayS(Asubject): #removes the holidays times from the subjects times
    week=Weekdays(weekdays)
    theMath=creatingTubles(Asubject)
    for x in range(len(Asubject)):
        for day in theMath:
                for days in week:
                    if day[0] == days:
                        theMath.remove(day)
    MathMath=theMath
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            for z in theMath:
                if sheet1.cell(row=x,column=y).value != None:
                    if sheet1.cell(row=x,column=sheet1.min_column).value == z[0] and sheet1.cell(row=sheet1.min_row,column=y).value == z[1]:
                        MathMath.remove(z)
    return MathMath
def PuttingValues():
    flag1=False
    flag2=False
    flag3=False
    flag4=False
    flag5=False
    flag6=False
    length=[len(removeHolidayS(math)),len(removeHolidayS(math2)),len(removeHolidayS(physics)),len(removeHolidayS(cs)),len(removeHolidayS(production)),len(removeHolidayS(elct))]
    dict={"math":removeHolidayS(math),"math2":removeHolidayS(math2),"physics":removeHolidayS(physics),"cs":removeHolidayS(cs),"production":removeHolidayS(production),"elct":removeHolidayS(elct)}
    yx = zip(list(dict.keys()), length)
    z=list(yx)
    z.sort()
    x_sorted = [dict for dict, length in z]
    x_sorted.reverse()
    i=0
    l=[]
    while i < len(emptySlot()):
        l+=[i]
        i+=1

    while flag1 == False:
        x=random.choice(l)
        for asd in dict[x_sorted[0]]:
            if str(asd) == str(emptySlot2()[x]):
                sheet[excel_style(emptySlot()[x][0],emptySlot()[x][1])]=x_sorted[0]+" lecture"
                flag1=True
                l.remove(x)
                break
            else:
                x=random.choice(l)

    while flag2 == False:
        y=random.choice(l)
        for fgh in dict[x_sorted[1]]:
            if str(fgh) == str(emptySlot2()[y]):
                sheet[excel_style(emptySlot()[y][0],emptySlot()[y][1])]=x_sorted[1]+" lecture"
                flag2=True
                l.remove(y)
                break
            else:
                y=random.choice(l)
    while flag3 == False:
        z=random.choice(l)
        for jkl in dict[x_sorted[2]]:
            if str(jkl) == str(emptySlot2()[z]):
                sheet[excel_style(emptySlot()[z][0],emptySlot()[z][1])]=x_sorted[2]+" lecture"
                flag3=True
                l.remove(z)
                break
            else:
                z=random.choice(l)
    while flag4 == False:
        w=random.choice(l)
        for zxc in dict[x_sorted[3]]:
            if str(zxc) == str(emptySlot2()[w]):
                sheet[excel_style(emptySlot()[w][0],emptySlot()[w][1])]=x_sorted[3]+" lecture"
                flag4=True
                l.remove(w)
                break
            else:
                w=random.choice(l)
    while flag5 == False:
        v=random.choice(l)
        for vbn in dict[x_sorted[4]]:
            if str(vbn) == str(emptySlot2()[v]):
                sheet[excel_style(emptySlot()[v][0],emptySlot()[v][1])]=x_sorted[4]+" lecture"
                flag5=True
                l.remove(v)
                break
            else:
                v=random.choice(l)
    while flag6 == False:
        p=random.choice(l)
        for iop in dict[x_sorted[5]]:
            if str(iop) == str(emptySlot2()[p]):
                sheet[excel_style(emptySlot()[p][0],emptySlot()[p][1])]=x_sorted[5]+" lecture"
                flag6=True
                l.remove(p)

                break
            else:
                p=random.choice(l)
    print("yes6")
    return flag1,flag2,flag3,flag4,flag5
PuttingValues()
def fullSlot():
    horizontal=sheet1.min_row+1
    vertical=sheet1.min_column+1
    slots=[]
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            if sheet1.cell(row=x,column =y).value != None:
                slots+=[(horizontal,vertical)]
            vertical+=1
        horizontal+=1
        vertical=sheet1.min_column+1
    return slots
i=0
l=[]
def copingOldSheet():
    i=0
    for x in range(sheet1.min_row,sheet1.max_row+1):
            for y in range(sheet1.min_column,sheet1.max_column+1):
                if sheet1.cell(row=x,column=y).value !=None:
                    sheet[excel_style(x,y)]=sheet1.cell(row=x,column=y).value
copingOldSheet()
book.save(r'C:\Users\Yousef\Documents\Python\appending.xlsx')
