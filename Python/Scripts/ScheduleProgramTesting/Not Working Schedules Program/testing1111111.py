import openpyxl
import random
from openpyxl import Workbook
book = Workbook()
sheet = book.active
fname = r'C:\Users\Yousef\Documents\Python\Schedule.xlsx'
wb = openpyxl.load_workbook(fname)
sheet1 = wb['Sheet1']
LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
weekdays=["saturday","sunday","monday","tuesday","wednesday","thursday"]
holidays=[]
math={"saturday":["1st"],"monday":["3rd","4th"],"wednesday":["2nd","4th"],"thursday":["4th"]}
physics={"saturday":["2nd","3rd"],"sunday":["4th"],"monday":["2nd"],"wednesday":["2nd","4th"]}
cs={"saturday":["3rd"],"sunday":["1st","2nd"],"tuesday":["3rd"],"thursday":["2nd","3rd"]}
production={"monday":["1st","2nd"],"wednesday":["1st","4th"]}
elct={"saturday":["2nd","4th"],"sunday":["3rd","4th"],"thursday":["4th"]}
subjects=["math","pyhsics","cs","prodcution","elct"]
def Weekdays(weekdays):
    holidays=[]
    flag=True
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            if sheet1.cell(row=x,column =y).value != None:
                flag = False
        if flag==True:
            holidays+=[sheet1.cell(row=x,column =sheet1.min_column).value]
            flag=False
        flag=True
    return holidays     #return a list of HOLIDAYS  #creates a list of the holidays
def creatingTubles(subject):
    tubles=[]
    for key,value in subject.items():
        for val in value:
            tubles+=[(key,val)]
    return tubles #creates tubles of times for a subject
def emptySlot2():
    horizontal=sheet1.min_row+1
    vertical=sheet1.min_column+1
    slots=[]
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            if len(Weekdays(weekdays)) ==2:
                if Weekdays(weekdays)[0] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[1] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(sheet1.cell(row=x,column =sheet1.min_column).value,sheet1.cell(row=sheet1.min_row,column =y).value)]
            elif len(Weekdays(weekdays)) ==1:
                if Weekdays(weekdays)[0] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(horizontal,vertical)]
            elif len(Weekdays(weekdays)) ==3:
                if Weekdays(weekdays)[0] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[1] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[2] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(horizontal,vertical)]
            vertical+=1
        horizontal+=1
        vertical=sheet1.min_column+1
    return slots    #getting empty slots from the Schedule before statring  #for the empty slots of the original schedule
def emptySlot():
    horizontal=sheet1.min_row+1
    vertical=sheet1.min_column+1
    slots=[]
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            if len(Weekdays(weekdays)) ==2:
                if Weekdays(weekdays)[0] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[1] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(horizontal,vertical)]
            elif len(Weekdays(weekdays)) ==1:
                if Weekdays(weekdays)[0] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(horizontal,vertical)]
            elif len(Weekdays(weekdays)) ==3:
                if Weekdays(weekdays)[0] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[1] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[2] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(horizontal,vertical)]
            vertical+=1
        horizontal+=1
        vertical=sheet1.min_column+1
    return slots
#def converSlotToCell():
    #names=[]
    #for x in emptySlot():
    #    names+=[excel_style(x[0],x[1])]
    #return names    #convert a slot to a name of cell using coordinates #convert cell coordinates to names
def coordinate(self):
        """This cell's coordinate (ex. 'A5')"""
        col = get_column_letter(self.column)
        return f"{col}{self.row}"   # not working yet
def excel_style(row, col):
    """ Convert given row and column number to an Excel-style cell name. """
    result = []
    while col:
        col, rem = divmod(col-1, 26)
        result[:0] = LETTERS [rem]
    return ''.join(result) + str(row)# the cell name system for excel   #getting the name of a cell in the format 'A1'  # the cell name system for excel   #getting the name of a cell in the format 'A1'
def removeHolidayS(Asubject): #removes the holidays times from the subjects times
    week=Weekdays(weekdays)
    theMath=creatingTubles(Asubject)
    for x in range(len(Asubject)):
        for day in theMath:
                for days in week:
                    if day[0] == days:
                        theMath.remove(day)
    return theMath
def PuttingValues():
    flag1=False
    flag2=False
    flag3=False
    flag4=False
    flag5=False
    while flag1 == False and flag2 == False and flag3 == False and flag4 == False and flag5 == False:
         i=0
         l=[]
         while i < len(emptySlot()):
             l+=[i]
             i+=1

         x=random.choice(l)
         for asd in removeHolidayS(math):
             if asd == emptySlot2(x):
                 sheet[emptySlot()[x]]="math lecture"
                 flag1=True
                 l.remove(x)
                 break
         y=random.choice(l)
         for fgh in removeHolidayS(physics):
             if fdh == emptySlot2(y):
                 sheet[emptySlot()[y]]="phy lecture"
                 flag2=True
                 l.remove(x)
                 break
         z=random.choice(l)
         for fgh in removeHolidayS(cs):
             if fdh == emptySlot2(z):
                 sheet[emptySlot()[z]]="cs lecture"
                 flag3=True
                 l.remove(x)
                 break
         w=random.choice(l)
         for fgh in removeHolidayS(production):
             if fdh == emptySlot2(w):
                 sheet[emptySlot()[w]]="production lecture"
                 flag4=True
                 l.remove(x)
                 break
         v=random.choice(l)
         for fgh in removeHolidayS(elct):
             if fdh == emptySlot2(v):
                 sheet[emptySlot()[v]]="elct lecture"
                 flag5=True
                 l.remove(x)
                 break



print(PuttingValues())


def appendingValues(sub):
    for x in range(sheet1.min_row,sheet1.max_row+1):
        for y in range(sheet1.min_column,sheet1.max_column+1):
            if sheet1.cell(row=x,column=y).value != None:
                sheet[excel_style(x,y)]=sheet1.cell(row=x,column=y).value
            else:
                for item in sub:
                    if item[0] == sheet1.cell(row=x,column=sheet1.min_column).value:
                        if item[1] == sheet1.cell(row=sheet1.min_row , column=y).value:
                            continue    #not working yet
#CellNames= converSlotToCell()

#LINE 105

book.save(r'C:\Users\Yousef\Documents\Python\appending.xlsx')























sheet[excel_style(emptySlot()[x][0],emptySlot()[x][1])]="math lecture"

def PuttingValues():
    flag1=False
    flag2=False
    flag3=False
    flag4=False
    flag5=False
    i=0
    l=[]
    while i < len(emptySlot()):
        l+=[i]
        i+=1

    while flag1 == False and flag2 == False and flag3 == False and flag4 == False and flag5 == False:
        x=random.choice(l)
        for asd in removeHolidayS(math):
            if str(asd) == str(emptySlot2()[x]):
                sheet[emptySlot()[x]]="math lecture"
                flag1=True
                l.remove(x)
                break
            else:
                x=random.choice(l)
        y=random.choice(l)
        for fgh in removeHolidayS(physics):
            if str(fgh) == str(emptySlot2()[y]):
                sheet[emptySlot()[y]]="physics lecture"
                flag2=True
                l.remove(y)
                break
            else:
                y=random.choice(l)
        z=random.choice(l)
        for jkl in removeHolidayS(cs):
            if str(jkl) == str(emptySlot2()[z]):
                sheet[emptySlot()[z]]="cs lecture"
                flag3=True
                l.remove(z)
                break
            else:
                z=random.choice(l)
        w=random.choice(l)
        for zxc in removeHolidayS(production):
            if str(zxc) == str(emptySlot2()[w]):
                sheet[emptySlot()[w]]="production lecture"
                flag4=True
                l.remove(w)
                break
            else:
                w=random.choice(l)
        v=random.choice(l)
        for vbn in removeHolidayS(elct):
            if str(vbn) == str(emptySlot2()[v]):
                sheet[emptySlot()[v]]="elct lecture"
                flag5=True
                l.remove(v)
                break
            else:
                v=random.choice(l)




















############latest one






def PuttingValues():
    flag1=False
    flag2=False
    flag3=False
    flag4=False
    flag5=False
    i=0
    l=[]
    while i < len(emptySlot()):
        l+=[i]
        i+=1

    while flag1 == False and flag2 == False and flag3 == False and flag4 == False and flag5 == False:
        x=random.choice(l)
        for asd in removeHolidayS(math):
            if str(asd) == str(emptySlot2()[x]):
                sheet[excel_style(emptySlot()[x][0],emptySlot()[x][1])]="math lecture"
                flag1=True
                l.remove(x)
                break
            else:
                x=random.choice(l)
        y=random.choice(l)
        for fgh in removeHolidayS(physics):
            if str(fgh) == str(emptySlot2()[y]):
                sheet[excel_style(emptySlot()[y][0],emptySlot()[y][1])]="physics lecture"
                flag2=True
                l.remove(y)
                break
            else:
                y=random.choice(l)
        z=random.choice(l)
        for jkl in removeHolidayS(cs):
            if str(jkl) == str(emptySlot2()[z]):
                sheet[excel_style(emptySlot()[z][0],emptySlot()[z][1])]="cs lecture"
                flag3=True
                l.remove(z)
                break
            else:
                z=random.choice(l)
        w=random.choice(l)
        for zxc in removeHolidayS(production):
            if str(zxc) == str(emptySlot2()[w]):
                sheet[excel_style(emptySlot()[w][0],emptySlot()[w][1])]="production lecture"
                flag4=True
                l.remove(w)
                break
            else:
                w=random.choice(l)
        v=random.choice(l)
        for vbn in removeHolidayS(elct):
            if str(vbn) == str(emptySlot2()[v]):
                sheet[emptySlot()[v]]="elct lecture"
                flag5=True
                l.remove(v)
                break
            else:
                v=random.choice(l)













def emptySlot3():
        horizontal=sheet1.min_row+1
        vertical=sheet1.min_column+1
        slots=[]
        for x in range(sheet1.min_row+1,sheet1.max_row+1):
            for y in range(sheet1.min_column+1,sheet1.max_column+1):
                if sheet1.cell(row=x,column =y).value != None:
                    full+=[(horizontal,vertical)]

                    vertical+=1
                    horizontal+=1
                    vertical=sheet1.min_column+1
                    return full
def copingOldSheet():
    i=0
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
            for y in range(sheet1.min_column+1,sheet1.max_column+1):
                if sheet1.cell(row=x,column=y).value !=None:
                    sheet[excel_style(emptySlot3()[i][0],emptySlot3()[i][1])]=sheet1.cell(row=x,column=y).value






















    while i < len(emptySlot()):
        l+=[i]
        i+=1

    while flag1 == False and flag2 == False and flag3 == False and flag4 == False and flag5 == False:

        x=random.choice(l)
        for asd in removeHolidayS(math):
            if str(asd) == str(emptySlot2()[x]):
                sheet[excel_style(emptySlot()[x][0],emptySlot()[x][1])]="math lecture"
                flag1=True
                l.remove(x)
                break
            else:
                x=random.choice(l)
    while flag2 == False:
        y=random.choice(l)
        for fgh in removeHolidayS(physics):
            if str(fgh) == str(emptySlot2()[y]):
                sheet[excel_style(emptySlot()[y][0],emptySlot()[y][1])]="physics lecture"
                flag2=True
                l.remove(y)
                break
            else:
                y=random.choice(l)
    while flag3 == False:
        z=random.choice(l)
        for jkl in removeHolidayS(cs):
            if str(jkl) == str(emptySlot2()[z]):
                sheet[excel_style(emptySlot()[z][0],emptySlot()[z][1])]="cs lecture"
                flag3=True
                l.remove(z)
                break
            else:
                z=random.choice(l)
    while flag4 == False:
        w=random.choice(l)
        for zxc in removeHolidayS(production):
            if str(zxc) == str(emptySlot2()[w]):
                sheet[excel_style(emptySlot()[w][0],emptySlot()[w][1])]="production lecture"
                flag4=True
                l.remove(w)
                break
            else:
                w=random.choice(l)
    while flag5 == False:
        v=random.choice(l)
        for vbn in removeHolidayS(elct):
            if str(vbn) == str(emptySlot2()[v]):
                sheet[excel_style(emptySlot()[v][0],emptySlot()[v][1])]="elct lecture"
                flag5=True
                l.remove(v)
                break
            else:
                v=random.choice(l)

    return flag1,flag2,flag3,flag4,flag5
