import itertools
import openpyxl
import random
from openpyxl import Workbook
book = Workbook()
sheet = book.active
fname = r'C:\Users\Yousef\Documents\Python\Schedule1.xlsx'
wb = openpyxl.load_workbook(fname)
sheet1 = wb['Sheet1']
LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
weekdays=["saturday","sunday","monday","tuesday","wednesday","thursday"]
holidays=[]
math={"saturday":["1st"],"monday":["3rd","4th"]}
math2={"wednesday":["2nd","4th"],"thursday":["4th"]}
physics={"saturday":["2nd","3rd"],"sunday":["2nd","4th"],"monday":["2nd"],"tuesday":["4th"],"wednesday":["2nd","4th"]}
cs={    "tuesday":["3rd"],"thursday":["2nd","3rd"]}
production={"monday":["1st","2nd"],"wednesday":["1st","4th"]}
elct={"saturday":["4th"],"sunday":["3rd","4th"],"thursday":["4th"]}
subjects=["math","pyhsics","cs","prodcution","elct"]
def Weekdays(weekdays):
    theWeekdays=[]
    holidays=[]
    flag=True
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            if sheet1.cell(row=x,column =y).value != None:
                flag = False
        if flag==True:
            holidays+=[sheet1.cell(row=x,column =sheet1.min_column).value]
            flag=False
        else:
            theWeekdays+=[sheet1.cell(row=x,column =sheet1.min_column).value]
        flag=True
    return theWeekdays,holidays #returns weekdays while the other returns holidays
def creatingTubles(subject):
    tubles=[]
    for key,value in subject.items():
        for val in value:
            tubles+=[(key,val)]
    return tubles #creates tubles of times for a subject
def emptySlot():
    horizontal=sheet1.min_row+1
    vertical=sheet1.min_column+1
    slots=[]
    slotsNames=[]
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            if len(Weekdays(weekdays)[1]) ==2:
                if Weekdays(weekdays)[1][0] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[1][1] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(horizontal,vertical)]
                            slotsNames+=[(sheet1.cell(row=x,column =sheet1.min_column).value,sheet1.cell(row=sheet1.min_row,column =y).value)]
            elif len(Weekdays(weekdays)[1]) ==1:
                if Weekdays(weekdays)[1][0] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(horizontal,vertical)]
                            slotsNames+=[(sheet1.cell(row=x,column =sheet1.min_column).value,sheet1.cell(row=sheet1.min_row,column =y).value)]
            elif len(Weekdays(weekdays)[1]) ==3:
                if Weekdays(weekdays)[1][0] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[1][1] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[1][2] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(horizontal,vertical)]
                            slotsNames+=[(sheet1.cell(row=x,column =sheet1.min_column).value,sheet1.cell(row=sheet1.min_row,column =y).value)]
            vertical+=1
        horizontal+=1
        vertical=sheet1.min_column+1
    return slotsNames,slots
def excel_style(row, col):
    """ Convert given row and column number to an Excel-style cell name. """
    result = []
    while col:
        col, rem = divmod(col-1, 26)
        result[:0] = LETTERS [rem]
    return ''.join(result) + str(row)# the cell name system for excel   #getting the name of a cell in the format 'A1'  # the cell name system for excel   #getting the name of a cell in the format 'A1'
def removeHolidayS(Asubject): #removes the holidays times from the subjects times
    week=Weekdays(weekdays)[1]
    theMath=creatingTubles(Asubject)
    for x in range(len(Asubject)):
        for day in theMath:
                for days in week:
                    if day[0] == days:
                        theMath.remove(day)
    MathMath=theMath
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            for z in theMath:
                if sheet1.cell(row=x,column=y).value != None:
                    if sheet1.cell(row=x,column=sheet1.min_column).value == z[0] and sheet1.cell(row=sheet1.min_row,column=y).value == z[1]:
                        MathMath.remove(z)
    return MathMath
def permutations():
    empty2=["math","math2","physics","cs","production","elct"]
    empty=list(itertools.permutations(empty2))
    return empty
def PuttingValues():
    for row in sheet['A1:G37']:
      for cell in row:
        cell.value = None
    options=permutations()
    subjects=["math","math2","physics","cs","production","elct"]
    dict={"math":removeHolidayS(math),"math2":removeHolidayS(math2),"physics":removeHolidayS(physics),"cs":removeHolidayS(cs),"production":removeHolidayS(production),"elct":removeHolidayS(elct)}
    empty=emptySlot()[0]
    emptyS=emptySlot()[1]
    empty_options=[]
    for option in options:
        empty_options+=[option]
    check={"math":False,"math2":False,"cs":False,"physics":False,"production":False,"elct":False}
    check_number={"math":0,"math2":0,"cs":0,"physics":0,"production":0,"elct":0}
    temp=empty
    check_number_temp=check_number
    i=0
    flag404=False
    counting_numbers={"math":0,"math2":0,"cs":0,"physics":0,"production":0,"elct":0}
    for key,values in dict.items():
        for the_subject_slot in values:
            for anOption in empty_options:
                for row in sheet['A1:G37']:
                    for cell in row:
                        cell.value = None
                empty=emptySlot()[0]
                emptyS=emptySlot()[1]
                check_number={"math":0,"math2":0,"cs":0,"physics":0,"production":0,"elct":0}
                for subject_name in anOption:
                    if check_number[subject_name] == 1:
                        break
                    if subject_name ==key:
                        for empty_slot in empty:
                            if the_subject_slot == empty_slot:
                                counter= empty.index(empty_slot)
                                sheet[excel_style(emptyS[counter][0],emptyS[counter][1])]= subject_name + " lecture"
                                check_number[subject_name]+=1
                                counting_numbers[subject_name]+=1
                                empty.remove(empty_slot)
                                emptyS.remove(emptyS[counter])
                                flag404=True
                                break
                        if check_number[subject_name] == 1:
                            break
                    else:
                        if check_number[subject_name] == 1:
                            break
                        for subject_slot in dict[subject_name]:
                            for empty_slot in empty:
                                if subject_slot == empty_slot:
                                    counter= empty.index(empty_slot)
                                    sheet[excel_style(emptyS[counter][0],emptyS[counter][1])]= subject_name + " lecture"
                                    check_number[subject_name]+=1
                                    counting_numbers[subject_name]+=1
                                    empty.remove(empty_slot)
                                    emptyS.remove(emptyS[counter])
                                    flag404=True
                                    break
                            if check_number[subject_name] == 1:
                                break
                flag=True
                for item,value in check_number.items():
                    if value !=1:
                        flag=False
                        break
                flag404=False
                flag_not_404=False
                if flag ==True:
                    i+=1
                    flag404=True
                    flag_not_404=True

            if flag_not_404 == True:
                break
        if flag_not_404 == True:
            break
    if flag404 == False:
        print(counting_numbers)
        print("sorry we couldn't do it")
        min=10000
        least_subject={}
        for key,value in counting_numbers.items():
            if value <min:
                least_subject={key:value}
                min=value
            elif value == min:
                least_subject[key]=value
        print("These subjects are causing us troubles in making your schedule: " + str(list(least_subject.keys()))+ ".Please add more time slots for them ")

PuttingValues()


def fullSlot():
    horizontal=sheet1.min_row+1
    vertical=sheet1.min_column+1
    slots=[]
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            if sheet1.cell(row=x,column =y).value != None:
                slots+=[(horizontal,vertical)]
            vertical+=1
        horizontal+=1
        vertical=sheet1.min_column+1
    return slots
i=0
l=[]
def copingOldSheet():
    i=0
    for x in range(sheet1.min_row,sheet1.max_row+1):
            for y in range(sheet1.min_column,sheet1.max_column+1):
                if sheet1.cell(row=x,column=y).value !=None:
                    sheet[excel_style(x,y)]=sheet1.cell(row=x,column=y).value
copingOldSheet()
book.save(r'C:\Users\Yousef\Documents\Python\appending.xlsx')
