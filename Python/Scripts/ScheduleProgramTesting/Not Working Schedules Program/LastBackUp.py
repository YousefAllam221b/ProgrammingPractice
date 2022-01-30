import itertools
import openpyxl
import random
from openpyxl import Workbook
book = Workbook()
sheet = book.active
fname = r'C:\Users\Yousef\Documents\Python\Schedule.xlsx'
wb = openpyxl.load_workbook(fname)
sheet1 = wb['Sheet1']
LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
weekdays=["saturday","sunday","monday","tuesday","wednesday","thursday"]
holidays=[]
math={"saturday":["1st"],"monday":["3rd","4th"]}
math2={"wednesday":["2nd","4th"],"thursday":["4th"]}
physics={"saturday":["2nd","3rd"],"sunday":["4th"],"monday":["2nd"],"wednesday":["2nd","4th"]}
cs={"tuesday":["3rd"],"thursday":["2nd","3rd"]}
production={"monday":["1st","2nd"],"wednesday":["1st","4th"]}
elct={"saturday":["2nd","4th"],"sunday":["3rd","4th"],"thursday":["4th"]}
subjects=["math","pyhsics","cs","prodcution","elct"]
def Weekdays(weekdays):
    holidays=[]
    flag=True
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            if sheet1.cell(row=x,column =y).value != None:
                flag = False
        if flag==True:
            holidays+=[sheet1.cell(row=x,column =sheet1.min_column).value]
            flag=False
        flag=True
    return holidays     #return a list of HOLIDAYS  #creates a list of the holidays
def Weekdays2(weekdays):
    holidays=[]
    flag=True
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            if sheet1.cell(row=x,column =y).value != None:
                flag = False
        if flag==True:
            weekdays.remove(sheet1.cell(row=x,column =sheet1.min_column).value)
            flag=False
        flag=True
    return weekdays #returns weekdays while the other returns holidays
def creatingTubles(subject):
    tubles=[]
    for key,value in subject.items():
        for val in value:
            tubles+=[(key,val)]
    return tubles #creates tubles of times for a subject
def emptySlot2():
    horizontal=sheet1.min_row+1
    vertical=sheet1.min_column+1
    slots=[]
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            if len(Weekdays(weekdays)) ==2:
                if Weekdays(weekdays)[0] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[1] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(sheet1.cell(row=x,column =sheet1.min_column).value,sheet1.cell(row=sheet1.min_row,column =y).value)]
            elif len(Weekdays(weekdays)) ==1:
                if Weekdays(weekdays)[0] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(horizontal,vertical)]
            elif len(Weekdays(weekdays)) ==3:
                if Weekdays(weekdays)[0] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[1] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[2] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(horizontal,vertical)]
            vertical+=1
        horizontal+=1
        vertical=sheet1.min_column+1
    return slots    #getting empty slots from the Schedule before statring  #for the empty slots of the original schedule
def emptySlot():
    horizontal=sheet1.min_row+1
    vertical=sheet1.min_column+1
    slots=[]
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            if len(Weekdays(weekdays)) ==2:
                if Weekdays(weekdays)[0] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[1] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(horizontal,vertical)]
            elif len(Weekdays(weekdays)) ==1:
                if Weekdays(weekdays)[0] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(horizontal,vertical)]
            elif len(Weekdays(weekdays)) ==3:
                if Weekdays(weekdays)[0] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[1] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[2] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(horizontal,vertical)]
            vertical+=1
        horizontal+=1
        vertical=sheet1.min_column+1
    return slots
def converSlotToCell():
    names=[]
    for x in emptySlot():
        names+=[excel_style(x[0],x[1])]
    return names    #convert a slot to a name of cell using coordinates #convert cell coordinates to names
def coordinate(self):
        """This cell's coordinate (ex. 'A5')"""
        col = get_column_letter(self.column)
        return f"{col}{self.row}"   # not working yet
def excel_style(row, col):
    """ Convert given row and column number to an Excel-style cell name. """
    result = []
    while col:
        col, rem = divmod(col-1, 26)
        result[:0] = LETTERS [rem]
    return ''.join(result) + str(row)# the cell name system for excel   #getting the name of a cell in the format 'A1'  # the cell name system for excel   #getting the name of a cell in the format 'A1'
def removeHolidayS(Asubject): #removes the holidays times from the subjects times
    week=Weekdays(weekdays)
    theMath=creatingTubles(Asubject)
    for x in range(len(Asubject)):
        for day in theMath:
                for days in week:
                    if day[0] == days:
                        theMath.remove(day)
    MathMath=theMath
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            for z in theMath:
                if sheet1.cell(row=x,column=y).value != None:
                    if sheet1.cell(row=x,column=sheet1.min_column).value == z[0] and sheet1.cell(row=sheet1.min_row,column=y).value == z[1]:
                        MathMath.remove(z)
    return MathMath
def checking(empty_par):
    flag10=False
    AList={}
    empty=empty_par
    subjects={"math":removeHolidayS(math),"math2":removeHolidayS(math2),"cs":removeHolidayS(cs),"physics":removeHolidayS(physics),"production":removeHolidayS(production),"elct":removeHolidayS(elct)}
    for x in subjects.keys():
        for y in subjects[x]:
            if y in empty:
                AList[x]=["True"]
                break
            else:
                AList[x]=["False"]
                break
    return AList
def permutations():
    empty2=["math","math2","physics","cs","production","elct"]
    empty=list(itertools.permutations(empty2))
    return empty

def PuttingValues():
    options=permutations()
    subjects=["math","math2","physics","cs","production","elct"]
    dict={"math":removeHolidayS(math),"math2":removeHolidayS(math2),"physics":removeHolidayS(physics),"cs":removeHolidayS(cs),"production":removeHolidayS(production),"elct":removeHolidayS(elct)}
    empty=emptySlot2()
    emptyS=emptySlot()
    i=0
    empty_options=[]
    for option in options:
        empty_options+=[option]
    flag404=False
    check={"math":False,"math2":False,"cs":False,"physics":False,"production":False,"elct":False}
    check_number={"math":0,"math2":0,"cs":0,"physics":0,"production":0,"elct":0}
    check_number_temp=check_number

    list=[]
    counter=0
    temp=empty
    tempS=emptyS
    check_temp=check
    counting=0
    for AnOption in empty_options:
        counting+=1
        appendedCheck=False
        optionCounter=0
        empty=temp
        emptyS=tempS
        check=check_temp
        check_number=check_number_temp
        for subject_name in AnOption:
            subject_numberCounter=0
            if check_number[subject_name] == 1:
                break
            for subject_slot in dict[subject_name]:
                counter=0
                appendedCheck=False
                subject_numberCounter=0
                for slot in empty:
                    if check_number[subject_name] == 1:
                        break
                    if subject_slot == slot:
                        sheet[excel_style(emptyS[counter][0],emptyS[counter][1])]=subject_name+" lecture"
                        empty.remove(slot)
                        emptyS.remove(emptyS[counter])
                        check_number[subject_name]+=1
                        check[subject_name]=True
                        break
                    counter+=1
            if check[subject_name]==1:
                break
        theFlag=True
        notappended=""
        for key,value in check.items():
            print(value)
            if value != True:
                theFlag=False
                notappended+=" " + key
                break
        if theFlag ==True:
            break
        print(check)
    print(check)
    print(check_number)
    print(counting)
PuttingValues()
def fullSlot():
    horizontal=sheet1.min_row+1
    vertical=sheet1.min_column+1
    slots=[]
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            if sheet1.cell(row=x,column =y).value != None:
                slots+=[(horizontal,vertical)]
            vertical+=1
        horizontal+=1
        vertical=sheet1.min_column+1
    return slots
i=0
l=[]
def copingOldSheet():
    i=0
    for x in range(sheet1.min_row,sheet1.max_row+1):
            for y in range(sheet1.min_column,sheet1.max_column+1):
                if sheet1.cell(row=x,column=y).value !=None:
                    sheet[excel_style(x,y)]=sheet1.cell(row=x,column=y).value
copingOldSheet()
book.save(r'C:\Users\Yousef\Documents\Python\appending.xlsx')
