import itertools
import openpyxl
import random
import os
from openpyxl import Workbook
book = Workbook()
sheet = book.active
fname = r'C:\Users\Yousef\Documents\Python\Schedule.xlsx'
wb = openpyxl.load_workbook(fname)
sheet1 = wb['Sheet1']
LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
weekdays=["saturday","sunday","monday","tuesday","wednesday","thursday"]
holidays=[]
math={"saturday":["1st","3rd","4th"],"monday":["3rd","4th"]}
math2={"wednesday":["2nd","4th"],"thursday":["1st","2nd","4th"]}
physics={"saturday":["2nd","3rd"],"sunday":["4th"],"monday":["2nd"],"wednesday":["2nd","4th"]}
cs={"saturday":["3rd"],"sunday":["1st","2nd"],"tuesday":["3rd"],"thursday":["2nd","3rd"]}
production={"monday":["1st","2nd"],"wednesday":["1st","4th"]}
elct={"saturday":["2nd","4th"],"sunday":["3rd","4th"],"thursday":["4th"]}
subjects=["math","pyhsics","cs","prodcution","elct"]

def Weekdays(weekdays):
    theWeekdays=[]
    holidays=[]
    flag=True
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            if sheet1.cell(row=x,column =y).value != None:
                flag = False
        if flag==True:
            holidays+=[sheet1.cell(row=x,column =sheet1.min_column).value]
            flag=False
        else:
            theWeekdays+=[sheet1.cell(row=x,column =sheet1.min_column).value]
        flag=True
    return theWeekdays,holidays #returns weekdays while the other returns holidays
def creatingTubles(subject):
    tubles=[]
    for key,value in subject.items():
        for val in value:
            tubles+=[(key,val)]
    return tubles #creates tubles of times for a subject
def emptySlot():
    horizontal=sheet1.min_row+1
    vertical=sheet1.min_column+1
    slots=[]
    slotsNames=[]
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            if len(Weekdays(weekdays)[1]) ==2:
                if Weekdays(weekdays)[1][0] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[1][1] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(horizontal,vertical)]
                            slotsNames+=[(sheet1.cell(row=x,column =sheet1.min_column).value,sheet1.cell(row=sheet1.min_row,column =y).value)]
            elif len(Weekdays(weekdays)[1]) ==1:
                if Weekdays(weekdays)[1][0] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(horizontal,vertical)]
                            slotsNames+=[(sheet1.cell(row=x,column =sheet1.min_column).value,sheet1.cell(row=sheet1.min_row,column =y).value)]
            elif len(Weekdays(weekdays)[1]) ==3:
                if Weekdays(weekdays)[1][0] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[1][1] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[1][2] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(horizontal,vertical)]
                            slotsNames+=[(sheet1.cell(row=x,column =sheet1.min_column).value,sheet1.cell(row=sheet1.min_row,column =y).value)]
            vertical+=1
        horizontal+=1
        vertical=sheet1.min_column+1
    return slotsNames,slots
def excel_style(row, col):
    """ Convert given row and column number to an Excel-style cell name. """
    result = []
    while col:
        col, rem = divmod(col-1, 26)
        result[:0] = LETTERS [rem]
    return ''.join(result) + str(row)# the cell name system for excel   #getting the name of a cell in the format 'A1'  # the cell name system for excel   #getting the name of a cell in the format 'A1'
def removeHolidayS(Asubject):
    daysoff=Weekdays(weekdays)[1]
    without_holidays={}
    for key,value in Asubject.items():
        if key not in daysoff:
            without_holidays[key]=value
    without_holidays=creatingTubles(without_holidays)
    last_list=[]
    for slot in without_holidays:
        if slot in emptySlot()[0]:
            last_list.append(slot)
    return last_list
def fullSlot():
    horizontal=sheet1.min_row+1
    vertical=sheet1.min_column+1
    slots=[]
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            if sheet1.cell(row=x,column =y).value != None:
                slots+=[(horizontal,vertical)]
            vertical+=1
        horizontal+=1
        vertical=sheet1.min_column+1
    return slots
def copingOldSheet(the_sheet):
    i=0
    for x in range(sheet1.min_row,sheet1.max_row+1):
            for y in range(sheet1.min_column,sheet1.max_column+1):
                if sheet1.cell(row=x,column=y).value !=None:
                    the_sheet[excel_style(x,y)]=sheet1.cell(row=x,column=y).value
def PuttingValues():
    sheet=book.active
    subjects=["math","math2","physics","cs","production","elct"]
    emptyS=emptySlot()[1]
    a = [removeHolidayS(math),removeHolidayS(math2),removeHolidayS(physics),removeHolidayS(cs),removeHolidayS(production),removeHolidayS(elct)]
    options=list(itertools.product(*a))
    list_without_duplicates=[]
    final_lists=[]
    if len(options) != 0:
        for option in options:
            list_without_duplicates=[]
            for i in option:
                if i not in list_without_duplicates:
                    list_without_duplicates.append(i)
            if len(list_without_duplicates) == len(option):
                final_lists.append(option)
        i=0
        tuples=[]
        bigger_list=[]
        for final_list in final_lists:
            tuples=[]
            i=0
            for slot in final_list:
                tuples.append((subjects[i],slot))
                i+=1
            bigger_list.append(tuples)

        o=1
        for item in bigger_list:
            for slot in item:
                the_index=emptySlot()[0].index(slot[1])
                sheet[excel_style(emptyS[the_index][0],emptyS[the_index][1])]= slot[0] + " lecture"
            copingOldSheet(sheet)
            if o<len(bigger_list):
                sheet=book.create_sheet("sheet " +str(o))
            o+=1
        book.save(r'C:\Users\Yousef\Documents\Python\appending.xlsx')
        file = r'C:\Users\Yousef\Documents\Python\appending.xlsx'
        os.startfile(file)
    else:
        print("no schedule found")
PuttingValues()
