import itertools
import openpyxl
import random
import os
from openpyxl import Workbook
book = Workbook()
sheet = book.active
#the next 3 lines are for reading the tutorials schedule.
fname = r'C:\Users\Yousef\Documents\Python\Schedules\Schedule.xlsx'
wb = openpyxl.load_workbook(fname)
sheet1 = wb['Sheet1']
LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
weekdays=["saturday","sunday","monday","tuesday","wednesday","thursday"]
holidays=[]
#The following variables are the time slots for each subject.
math={"saturday":["1st","3rd","4th"],"monday":["3rd","4th"]}
math2={"wednesday":["2nd","4th"],"thursday":["1st","2nd","4th"]}
physics={"saturday":["2nd","3rd"],"sunday":["4th"],"monday":["2nd"],"wednesday":["2nd","4th"]}
cs={"saturday":["3rd"],"sunday":["1st","2nd"],"tuesday":["3rd"],"thursday":["2nd","3rd"]}
production={"monday":["1st","2nd"],"wednesday":["1st","4th"]}
elct={"saturday":["2nd","4th"],"sunday":["3rd","4th"],"thursday":["4th"]}

#Weekdays() function take a single parameter which is the weekdays variable above
#and returns a list with the weekdays and other list with holidays.
def Weekdays(weekdays):
    theWeekdays=[]
    holidays=[]
    flag=True
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            if sheet1.cell(row=x,column =y).value != None:
                flag = False
        if flag==True:
            holidays+=[sheet1.cell(row=x,column =sheet1.min_column).value]
            flag=False
        else:
            theWeekdays+=[sheet1.cell(row=x,column =sheet1.min_column).value]
        flag=True
    return theWeekdays,holidays #returns weekdays while the other returns holidays
#creating_tuples function take the dictionary of a subject
#and creates a list of tuples where each tuple is a time slot
#in the form (day,period).
def creating_tuples(subject):
    tubles=[]
    for key,value in subject.items():
        for val in value:
            tubles+=[(key,val)]
    return tubles #creates tubles of times for a subject
#emptySlot() function creates a list with all empt slots in the tutorials schedule
#and another list of their position in the schedule in the form (row,column).
def emptySlot():
    horizontal=sheet1.min_row+1
    vertical=sheet1.min_column+1
    slots=[]
    slotsNames=[]
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            if len(Weekdays(weekdays)[1]) ==2:
                if Weekdays(weekdays)[1][0] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[1][1] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(horizontal,vertical)]
                            slotsNames+=[(sheet1.cell(row=x,column =sheet1.min_column).value,sheet1.cell(row=sheet1.min_row,column =y).value)]
            elif len(Weekdays(weekdays)[1]) ==1:
                if Weekdays(weekdays)[1][0] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(horizontal,vertical)]
                            slotsNames+=[(sheet1.cell(row=x,column =sheet1.min_column).value,sheet1.cell(row=sheet1.min_row,column =y).value)]
            elif len(Weekdays(weekdays)[1]) ==3:
                if Weekdays(weekdays)[1][0] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[1][1] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[1][2] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(horizontal,vertical)]
                            slotsNames+=[(sheet1.cell(row=x,column =sheet1.min_column).value,sheet1.cell(row=sheet1.min_row,column =y).value)]
            vertical+=1
        horizontal+=1
        vertical=sheet1.min_column+1
    return slotsNames,slots
#excel_style() function is used to get the name of a cell in the excel sheet.
#It takes two parameters the row number and column number where the cell lies.
def excel_style(row, col):
    """ Convert given row and column number to an Excel-style cell name. """
    result = []
    while col:
        col, rem = divmod(col-1, 26)
        result[:0] = LETTERS [rem]
    return ''.join(result) + str(row)# the cell name system for excel   #getting the name of a cell in the format 'A1'  # the cell name system for excel   #getting the name of a cell in the format 'A1'
#the filter_slots function is used to filter a subject from slots that cant be used
#as slots which are in a dayoff or a slot which is already occupied with a tutorial.
def filter_slots(Asubject):
    #formly removeHolidayS()
    daysoff=Weekdays(weekdays)[1]
    without_holidays={}
    for key,value in Asubject.items():
        if key not in daysoff:
            without_holidays[key]=value
    without_holidays=creating_tuples(without_holidays)
    last_list=[]
    for slot in without_holidays:
        if slot in emptySlot()[0]:
            last_list.append(slot)
    return last_list
#copingOldSheet() function is used to copy the tutorials schedule in the new sheets of the possible schedules.
def copingOldSheet(the_sheet):
    i=0
    for x in range(sheet1.min_row,sheet1.max_row+1):
            for y in range(sheet1.min_column,sheet1.max_column+1):
                if sheet1.cell(row=x,column=y).value !=None:
                    the_sheet[excel_style(x,y)]=sheet1.cell(row=x,column=y).value
def PuttingValues():
    sheet=book.active
    subjects=["Math","Math2","Physics","Csen","Production","Elct"]
    emptyS=emptySlot()[1]
    subjects_slots = [filter_slots(math),filter_slots(math2),filter_slots(physics),filter_slots(cs),filter_slots(production),filter_slots(elct)]
    #the next function  generates a list of all possible combinations.
    options=list(itertools.product(*subjects_slots))
    list_without_duplicates=[]
    final_lists=[]
    #the condition is used to know if there is possible lists or not
    #as if a subject doesnt have any avaliable slot after using filter_slots() function,
    #there will be no possible lists and this will output to there is no avaliable schedule.
    if len(options) != 0:
        #The next loop filter the options list from any list contains two or more identical time slot.
        for option in options:
            list_without_duplicates=[]
            for i in option:
                if i not in list_without_duplicates:
                    list_without_duplicates.append(i)
            if len(list_without_duplicates) == len(option):
                final_lists.append(option)
        i=0
        tuples=[]
        bigger_list=[]
        #The next loop join the subject name with it's time slot as it takes each slot in order
        #and join it with the a subject name of the same order.
        for final_list in final_lists:
            tuples=[]
            i=0
            for slot in final_list:
                tuples.append((subjects[i],slot))
                i+=1
            bigger_list.append(tuples)
        o=2
        #the next loop takes the lists of the working schedules and put each list in a sheet
        for alist in bigger_list:
            for slot in alist:
                counter=emptySlot()[0].index(slot[1])
                sheet[excel_style(emptyS[counter][0],emptyS[counter][1])]= slot[0] + " lecture"
            copingOldSheet(sheet)
            if o <= len(bigger_list):
                sheet=book.create_sheet('Sheet '+ str(o))
            o+=1
        #The next 5 lines before the else part are for changing the name of the first sheet
        #from Sheet to Sheet 1 and also opens the excel file.
        sheet = book['Sheet']
        sheet.title = 'Sheet 1'
        book.save(r'C:\Users\Yousef\Documents\Python\appending.xlsx')
        # file = r'C:\Users\Yousef\Documents\Python\appending.xlsx'
        # os.startfile(file)
    else:
        #the next print statement is for the user to know that there is no possible schedule with
        #the given input.
        print("No possible schedule could be formed with the input you gave.")
        print(options)
#Taking_Schedules() function creates a list of all the schedules generated from the program.
#the list is in the form of a list contains schedules ,each contains lists of days.
PuttingValues()
fname = r'C:\Users\Yousef\Documents\Python\appending.xlsx'
wb2 = openpyxl.load_workbook(fname)
sheet1=wb2["Sheet 1"]
def Taking_Schedules():
    holidays=Weekdays(weekdays)[1]
    sheets=wb2.sheetnames
    all_data=[]
    sheet_data=[]
    day=[]
    week=[]
    schedules=[]
    #the following loop takes the data from every sheet in form of a list and save it in a bigger list.
    for sheet in sheets:
        aSheet=book[sheet]
        sheet_data=[]
        week=[]
        for x in range(aSheet.min_row+1,aSheet.max_row+1):
            day=[]
            if aSheet.cell(row=x,column =aSheet.min_column).value in holidays:
                continue
            for y in range(aSheet.min_column+1,aSheet.max_column+1):
                    day+=[aSheet.cell(row=x,column =y).value]
            week+=[day]
        all_data+=[week]
    return all_data
#First_or_Fifth() return list of schedules with most free first slots or fifth slots.
print(Taking_Schedules())
def First_or_Fifth(needed):
    data=needed
    with_gaps=[]
    no_first=[]
    no_fifth=[]
    i=0
    j=0
    k=0
    max=0
    max2=0
    # for the next loop,in the first condition , it finds the schedules which have the most free first slots.
    #in the second condition , if finds the schedules which have the most free fifth slots.
    for schedule in data:
        j=0
        k=0
        for day in schedule:
                    if day[0] == None:
                        j+=1
                    if day[4] ==None:
                        k+=1
        if j == max :
            no_first+=[schedule]
        elif j> max:
            max=j
            no_first=[schedule]
        if k == max2 :
            no_fifth+=[schedule]
        elif k> max2:
            max2=k
            no_fifth=[schedule]

    return no_first,no_fifth
#gaps() function return two list: one with schedules with gapes in them and the other is withoutgaps.
def gaps(needed):
    i=0
    schedules=needed
    with_gaps=[]
    flag=False
    for week in schedules:
        for day in week:
            i=0
            flag=False
            a=day[1:4]
            if None in a:
                if day[0] != None and day[4] != None:
                    with_gaps+=[week]
                    break
                elif day[0] != None and day[4] == None:
                    i=a.index(None)
                    while i <2:
                        i+=1
                        if a[i] != None:
                            with_gaps+=[week]
                            flag=True
                            break
                    if flag == True:
                        break
                elif day[4] != None and day[0] == None:
                    i=a.index(None)
                    while i > 0:
                        i-=1
                        if a[i] != None:
                            with_gaps+=[week]
                            flag=True
                            break
                    if flag == True:
                        break
                elif day[0] == None and day[4] == None:
                    i=a.index(None)
                    if i ==1:
                        if day[1] != None and day[3] != None:
                            with_gaps+=[week]
                            flag=True
                            break    #50914=25650 +1026 +23306+932
    without_gaps=[]
    for item in schedules:
        if item not in with_gaps:
            without_gaps.append(item)
    return with_gaps , without_gaps
answer=input("Can we help you more with choosing the best schedule for you?(if it exits)(please answer with Y/N) ")
if answer.lower() == "y" or answer.lower() == "yes":
    next_answer=input("Do you like a schedule with gaps between slots or with no gaps?(please enter 0 for gaps and 1 for no gaps) ")
    after_answer=input("Do you like a schedule with the most free first slots or with the most free fifth slots?(please enter 0 for free firsts and 1 for free fifths) ")

def asking():
    x=2
    y=2
    if answer.lower() == "y" or answer.lower() == "yes":
        if next_answer =="0" or next_answer.lower().replace(" ", "") =="gaps":
            x=0
        else:
            x=1
        if after_answer =="0" or next_answer.lower().replace(" ", "") =="freefirst":
            y=0
        else:
            y=1
        out=gaps(First_or_Fifth(Taking_Schedules())[x])[y]


        flag1=True
        return flag1,out
    else:
        file = r'C:\Users\Yousef\Documents\Python\appending.xlsx'
        os.startfile(file)
ab=Workbook()
she=ab.active
#the next block of code(the body of the if) is not working as the two loops works only with the big list not the matrix from Taking_Schedules()
if asking()[0] == True:
    daysoff=Weekdays(weekdays)[1]
    indices=[]
    for item in daysoff:
        indices+=[weekdays.index(item)+2]
    o=2
    k=1
    my_lists=asking()[1]
    for week in my_lists:
        x=2
        for day in week:
            y=2
            while x in indices:

                x+=1
            for slot in day:
                she[excel_style(x,y)]= slot
                y+=1
            x+=1
        if k<len(my_lists):
            for x in range(sheet1.min_row,sheet1.max_row+1):
                for y in range(sheet1.min_column,sheet1.max_column+1):
                    if sheet1.cell(row=x,column=y).value !=None and "tut" not in sheet1.cell(row=x,column=y).value and "lecture" not in sheet1.cell(row=x,column=y).value:
                        she[excel_style(x,y)]=sheet1.cell(row=x,column=y).value
            she=ab.create_sheet("Sheet " + str(o))

        o+=1
        k+=1
    i=0
for x in range(sheet1.min_row,sheet1.max_row+1):
    for y in range(sheet1.min_column,sheet1.max_column+1):
        if sheet1.cell(row=x,column=y).value !=None and "tut" not in sheet1.cell(row=x,column=y).value and "lecture" not in sheet1.cell(row=x,column=y).value:
            she[excel_style(x,y)]=sheet1.cell(row=x,column=y).value

sheet = ab['Sheet']
sheet.title = 'Sheet 1'
ab.save(r'C:\Users\Yousef\Documents\Python\appending.xlsx')
