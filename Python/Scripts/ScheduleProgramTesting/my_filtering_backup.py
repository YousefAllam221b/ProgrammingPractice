import openpyxl
fname = r'C:\Users\Yousef\Documents\Python\appending.xlsx'
wb = openpyxl.load_workbook(fname)
weekdays=["saturday","sunday","monday","tuesday","wednesday","thursday"]
sheet1=wb["Sheet 1"]
#the same Weekdays() from the program.
def Weekdays(weekdays):
    theWeekdays=[]
    holidays=[]
    flag=True
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            if sheet1.cell(row=x,column =y).value != None:
                flag = False
        if flag==True:
            holidays+=[sheet1.cell(row=x,column =sheet1.min_column).value]
            flag=False
        else:
            theWeekdays+=[sheet1.cell(row=x,column =sheet1.min_column).value]
        flag=True
    return theWeekdays,holidays
#Taking_Schedules() function creates a list of all the schedules generated from the program.
#the list is in the form of a list contains schedules ,each contains lists of days.
def Taking_Schedules():
    holidays=Weekdays(weekdays)[1]
    sheets=wb.sheetnames
    all_data=[]
    sheet_data=[]
    day=[]
    week=[]
    schedules=[]
    #the following loop takes the data from every sheet in form of a list and save it in a bigger list.
    for sheet in sheets:
        aSheet=wb[sheet]
        sheet_data=[]
        week=[]
        for x in range(aSheet.min_row+1,aSheet.max_row+1):
            day=[]
            if aSheet.cell(row=x,column =aSheet.min_column).value in holidays:
                continue
            for y in range(aSheet.min_column+1,aSheet.max_column+1):
                    day+=[aSheet.cell(row=x,column =y).value]
            week+=[day]
        all_data+=[week]
    return all_data
#First_or_Fifth() return list of schedules with most free first slots or fifth slots.
def First_or_Fifth(needed):
    data=needed
    with_gaps=[]
    no_first=[]
    no_fifth=[]
    i=0
    j=0
    k=0
    max=0
    max2=0
    # for the next loop,in the first condition , it finds the schedules which have the most free first slots.
    #in the second condition , if finds the schedules which have the most free fifth slots.
    for schedule in data:
        j=0
        k=0
        for day in schedule:
                    if day[0] == None:
                        j+=1
                    if day[4] ==None:
                        k+=1
        if j == max :
            no_first+=[schedule]
        elif j> max:
            max=j
            no_first=[schedule]
        if k == max2 :
            no_fifth+=[schedule]
        elif k> max2:
            max2=k
            no_fifth=[schedule]

    return no_first,no_fifth
#gaps() function return two list: one with schedules with gapes in them and the other is withoutgaps.
def gaps(needed):
    i=0
    schedules=needed
    with_gaps=[]
    flag=False
    for week in schedules:
        for day in week:
            i=0
            flag=False
            a=day[1:4]
            if None in a:
                if day[0] != None and day[4] != None:
                    with_gaps+=[week]
                    break
                elif day[0] != None and day[4] == None:
                    i=a.index(None)
                    while i <2:
                        i+=1
                        if a[i] != None:
                            with_gaps+=[week]
                            flag=True
                            break
                    if flag == True:
                        break
                elif day[4] != None and day[0] == None:
                    i=a.index(None)
                    while i > 0:
                        i-=1
                        if a[i] != None:
                            with_gaps+=[week]
                            flag=True
                            break
                    if flag == True:
                        break
                elif day[0] == None and day[4] == None:
                    i=a.index(None)
                    if i ==1:
                        if day[1] != None and day[3] != None:
                            with_gaps+=[week]
                            flag=True
                            break    #50914=25650 +1026 +23306+932
    without_gaps=[]
    for item in schedules:
        if item not in with_gaps:
            without_gaps.append(item)
    return with_gaps , without_gaps
