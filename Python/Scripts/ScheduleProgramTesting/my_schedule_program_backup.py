import itertools
import openpyxl
import random
import os
from openpyxl import Workbook
book = Workbook()
sheet = book.active
#the next 3 lines are for reading the tutorials schedule.
fname = r'C:\Users\Yousef\Documents\Python\Schedules\Schedule.xlsx'
wb = openpyxl.load_workbook(fname)
sheet1 = wb['Sheet1']
LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
weekdays=["saturday","sunday","monday","tuesday","wednesday","thursday"]
holidays=[]
#The following variables are the time slots for each subject.
math={"saturday":["1st","3rd","4th"],"monday":["3rd","4th"]}
math2={"wednesday":["2nd","4th"],"thursday":["1st","2nd","4th"]}
physics={"saturday":["2nd","3rd"],"sunday":["4th"],"monday":["2nd"],"wednesday":["2nd","4th"]}
cs={"saturday":["3rd"],"sunday":["1st","2nd"],"tuesday":["3rd"],"thursday":["2nd","3rd"]}
production={"monday":["1st","2nd"],"wednesday":["1st","4th"]}
elct={"saturday":["2nd","4th"],"sunday":["3rd","4th"],"thursday":["4th"]}
#Weekdays() function take a single parameter which is the weekdays variable above
#and returns a list with the weekdays and other list with holidays.
def Weekdays(weekdays):
    theWeekdays=[]
    holidays=[]
    flag=True
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            if sheet1.cell(row=x,column =y).value != None:
                flag = False
        if flag==True:
            holidays+=[sheet1.cell(row=x,column =sheet1.min_column).value]
            flag=False
        else:
            theWeekdays+=[sheet1.cell(row=x,column =sheet1.min_column).value]
        flag=True
    return theWeekdays,holidays #returns weekdays while the other returns holidays
#creating_tuples function take the dictionary of a subject
#and creates a list of tuples where each tuple is a time slot
#in the form (day,period).
def creating_tuples(subject):
    tubles=[]
    for key,value in subject.items():
        for val in value:
            tubles+=[(key,val)]
    return tubles #creates tubles of times for a subject
#emptySlot() function creates a list with all empt slots in the tutorials schedule
#and another list of their position in the schedule in the form (row,column).
def emptySlot():
    horizontal=sheet1.min_row+1
    vertical=sheet1.min_column+1
    slots=[]
    slotsNames=[]
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            if len(Weekdays(weekdays)[1]) ==2:
                if Weekdays(weekdays)[1][0] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[1][1] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(horizontal,vertical)]
                            slotsNames+=[(sheet1.cell(row=x,column =sheet1.min_column).value,sheet1.cell(row=sheet1.min_row,column =y).value)]
            elif len(Weekdays(weekdays)[1]) ==1:
                if Weekdays(weekdays)[1][0] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(horizontal,vertical)]
                            slotsNames+=[(sheet1.cell(row=x,column =sheet1.min_column).value,sheet1.cell(row=sheet1.min_row,column =y).value)]
            elif len(Weekdays(weekdays)[1]) ==3:
                if Weekdays(weekdays)[1][0] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[1][1] != sheet1.cell(row=x,column=sheet1.min_column).value and Weekdays(weekdays)[1][2] != sheet1.cell(row=x,column=sheet1.min_column).value:
                        if sheet1.cell(row=x,column =y).value == None:
                            slots+=[(horizontal,vertical)]
                            slotsNames+=[(sheet1.cell(row=x,column =sheet1.min_column).value,sheet1.cell(row=sheet1.min_row,column =y).value)]
            vertical+=1
        horizontal+=1
        vertical=sheet1.min_column+1
    return slotsNames,slots
#excel_style() function is used to get the name of a cell in the excel sheet.
#It takes two parameters the row number and column number where the cell lies.
def excel_style(row, col):
    """ Convert given row and column number to an Excel-style cell name. """
    result = []
    while col:
        col, rem = divmod(col-1, 26)
        result[:0] = LETTERS [rem]
    return ''.join(result) + str(row)# the cell name system for excel   #getting the name of a cell in the format 'A1'  # the cell name system for excel   #getting the name of a cell in the format 'A1'
#the filter_slots function is used to filter a subject from slots that cant be used
#as slots which are in a dayoff or a slot which is already occupied with a tutorial.
def filter_slots(Asubject):
    #formly removeHolidayS()
    daysoff=Weekdays(weekdays)[1]
    without_holidays={}
    for key,value in Asubject.items():
        if key not in daysoff:
            without_holidays[key]=value
    without_holidays=creating_tuples(without_holidays)
    last_list=[]
    for slot in without_holidays:
        if slot in emptySlot()[0]:
            last_list.append(slot)
    return last_list
#copingOldSheet() function is used to copy the tutorials schedule in the new sheets of the possible schedules.
def copingOldSheet(the_sheet):
    i=0
    for x in range(sheet1.min_row,sheet1.max_row+1):
            for y in range(sheet1.min_column,sheet1.max_column+1):
                if sheet1.cell(row=x,column=y).value !=None:
                    the_sheet[excel_style(x,y)]=sheet1.cell(row=x,column=y).value
def PuttingValues():
    sheet=book.active
    subjects=["Math","Math2","Physics","Csen","Production","Elct"]
    emptyS=emptySlot()[1]
    subjects_slots = [filter_slots(math),filter_slots(math2),filter_slots(physics),filter_slots(cs),filter_slots(production),filter_slots(elct)]
    #the next function  generates a list of all possible combinations.
    options=list(itertools.product(*subjects_slots))
    list_without_duplicates=[]
    final_lists=[]
    #the condition is used to know if there is possible lists or not
    #as if a subject doesnt have any avaliable slot after using filter_slots() function,
    #there will be no possible lists and this will output to there is no avaliable schedule.
    if len(options) != 0:
        #The next loop filter the options list from any list contains two or more identical time slot.
        for option in options:
            list_without_duplicates=[]
            for i in option:
                if i not in list_without_duplicates:
                    list_without_duplicates.append(i)
            if len(list_without_duplicates) == len(option):
                final_lists.append(option)
        i=0
        tuples=[]
        bigger_list=[]
        #The next loop join the subject name with it's time slot as it takes each slot in order
        #and join it with the a subject name of the same order.
        for final_list in final_lists:
            tuples=[]
            i=0
            for slot in final_list:
                tuples.append((subjects[i],slot))
                i+=1
            bigger_list.append(tuples)
        o=2
        #the next loop takes the lists of the working schedules and put each list in a sheet
        for alist in bigger_list:
            for slot in alist:
                counter=emptySlot()[0].index(slot[1])
                sheet[excel_style(emptyS[counter][0],emptyS[counter][1])]= slot[0] + " lecture"
            copingOldSheet(sheet)
            if o <= len(bigger_list):
                sheet=book.create_sheet('Sheet '+ str(o))
            o+=1
        #The next 5 lines before the else part are for changing the name of the first sheet
        #from Sheet to Sheet 1 and also opens the excel file.
        sheet = book['Sheet']
        sheet.title = 'Sheet 1'
        book.save(r'C:\Users\Yousef\Documents\Python\appending.xlsx')
        file = r'C:\Users\Yousef\Documents\Python\appending.xlsx'
        os.startfile(file)
    else:
        #the next print statement is for the user to know that there is no possible schedule with
        #the given input.
        print("No possible schedule could be formed with the input you gave.")

PuttingValues()
